import sys
import json
import os
import openpyxl
import errno
import csv
import Module.logger
import xml.etree.ElementTree as ET
import PyPDF2
import Module.cfg
import imaplib
import email
import datetime

def ReadDataFromJsonFile(type,value):
    #print("reading data from json file")
    ## Read Json file from Config directory
    cwd = os.getcwd()
    pcwd = "\\".join(cwd.split('\\')[:-1])
    confdir = cwd + "\\Config"
    try:
        dirExists = os.path.exists(confdir)
        if not dirExists:
            confdir = pcwd + "\\Config"
    except OSError as exception:
            if exception.errno != errno.EEXIST:
                raise

    #confdir = pcwd + "\\Config"
    #confdir = pcwd + "\\Config"
    json_data = open(confdir+"\\AutomationToolConfig.json")
    data = json.load(json_data)
   # print(data)
    try:
      return data[type][value]
    except:
      return None

def ReadDataFromLoginFile(type,value):
    #print("reading data from json file")
    ## Read Json file from Config directory
    cwd = os.getcwd()
    pcwd = "\\".join(cwd.split('\\')[:-1])
    confdir = cwd + "\\Config"
    try:
        dirExists = os.path.exists(confdir)
        if not dirExists:
            confdir = pcwd + "\\Config"
    except OSError as exception:
            if exception.errno != errno.EEXIST:
                raise

    #confdir = pcwd + "\\Config"
    #confdir = pcwd + "\\Config"
    json_data = open(confdir+"\\Login.json")
    data = json.load(json_data)
   # print(data)
    try:
      return data[type][value]
    except:
      return None

def ReadDataFromRepo(type, name, value):
    get_input_type = ReadDataFromJsonFile("tool", "configfile")
    if get_input_type == "xls":
        strColValues = type+"|"+name
        try:
            return fnReadDataFromExcel("LogicalValues.xlsx", "LogicalNames", "Type|Name", strColValues, value)
        except:
            return None
    else:
        cwd = os.getcwd()
        pcwd = "\\".join(cwd.split('\\')[:-1])
        confdir = cwd + "\\Config"
        try:
            dirExists = os.path.exists(confdir)
            if not dirExists:
                confdir = pcwd + "\\Config"
        except OSError as exception:
            if exception.errno != errno.EEXIST:
                raise
        repo_file = open(confdir+"\\Repository.json")
        repo_data = json.load(repo_file)
        try:
            return repo_data[type][name][value]
        except:
            return None

def fnReadDataFromExcel(strExcelName,strSheetName,strUniqueColNames, strUniqueColValues, strValue):
        cwd = os.getcwd()
        pcwd = "\\".join(cwd.split('\\')[:-1])
        confdir = cwd + "\\Config"
        try:
            dirExists = os.path.exists(confdir)
            if not dirExists:
                confdir = pcwd + "\\Config"
        except OSError as exception:
            if exception.errno != errno.EEXIST:
                raise
        wb = openpyxl.load_workbook(confdir+"\\" + strExcelName)
        sheet = wb.get_sheet_by_name(strSheetName)
        maxrows = sheet.max_row
        maxcols = sheet.max_column
        blnfound = False
        blnList = False

        ColValuesList = strUniqueColValues.split('|')
        if '|'  in strUniqueColValues:
         blnList = True

        for r in range(2,maxrows+1):
            counter = 0
            for c in range(1,maxcols+1):
                for i in range(0,len(ColValuesList)):
                    if sheet.cell(row=r,column=c).value == ColValuesList[i]:
                        counter+=1
                        if counter == len(ColValuesList) and blnList == True:
                            UniqueRowNo = r
                            #print("Unique Row number is : "+str(UniqueRowNo))
                            blnfound = True
                            break
                        if blnList!=True:
                            if counter == 1:
                                UniqueRowNo = r
                                #print("Unique Row number is : " + str(UniqueRowNo))
                                blnfound = True
                                break
        if blnfound :
            for getcolnum in range(1, maxcols + 1):
                if sheet.cell(row=1, column=getcolnum).value == strValue:
                    #print("Uniquecolvalue = " + str(getcolnum))
                    break
            c = sheet.cell(row=UniqueRowNo, column=getcolnum).value
            #print("The Value of " + strValue + " based on Unique Column and Row Values is : " + str(c))
            return c



def CheckIfDefinedElementExistInRepo(type,name):
    idtype = ReadDataFromRepo(type,name,"locator")
    idvalue = ReadDataFromRepo(type,name,"locatorvalue")
    return idtype,idvalue

def getDataForDynamicAlgo(type):
    locator = ReadDataFromJsonFile(type,"locator")
    locatorvalue = ReadDataFromJsonFile(type,"locatorvalue")
    return locator,locatorvalue



def readTestData(value):
    cwd = os.getcwd()
    pcwd = "\\".join(cwd.split('\\')[:-1])
    confdir = cwd + "\\TestData"
    try:
        dirExists = os.path.exists(confdir)
        if not dirExists:
            confdir = pcwd + "\\TestData"
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise

    json_data = open(confdir + "\\TestData.json")
    data = json.load(json_data)
    if "Row" in value:
        value = value.split("'")[1]
    if "USE_" in value:
        all_values = value.split("_")
        type = all_values[1]
        print("Type is"+type)
        value = all_values[2]
        print("Value is" + value)
        try:
            return data[type][value]
        except:
            return value
    else:
        return value

def ReadDataFromMobileConfig(type,value):
    cwd = os.getcwd()
    pcwd = "\\".join(cwd.split('\\')[:-1])
    confdir = cwd + "\\Config"
    try:
        dirExists = os.path.exists(confdir)
        if not dirExists:
            confdir = pcwd + "\\Config"
    except OSError as exception:
            if exception.errno != errno.EEXIST:
                raise

    json_data = open(confdir+"\\MobileConfig.json")
    data = json.load(json_data)
   # print(data)
    try:
      return data[type][value]
    except:
      return None

def getMobileDataForDynamicAlgo(type):
    locator = ReadDataFromMobileConfig(type,"locator")
    locatorvalue = ReadDataFromMobileConfig(type,"locatorvalue")
    return locator,locatorvalue


def getFileTransferPath(filename):
    cwd = os.getcwd()
    pcwd = "\\".join(cwd.split('\\')[:-1])
    filePath = cwd + "\\fileTransfer\\"+filename
    try:
        dirExists = os.path.exists(filePath)
        if not dirExists:
            filePath = pcwd + "\\fileTransfer\\"+filename
            return filePath
        else:
            return filePath
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise

def defaultDownloadDirectory():
    cwd = os.getcwd()
    pcwd = "\\".join(cwd.split('\\')[:-1])
    downloadDir = cwd + "\\fileTransfer\\"
    try:
        dirExists = os.path.exists(downloadDir)
        if not dirExists:
            downloadDir = pcwd + "\\fileTransfer\\"
            return downloadDir
        else:
            return downloadDir
    except OSError as exception:
        if exception.errno != errno.EEXIST:
            raise

def get_latest_file():
    defaultDir = defaultDownloadDirectory()
    paths = [os.path.join(defaultDir, fname) for fname in os.listdir(defaultDir)]
    time_sorted_list = sorted(paths, key=os.path.getctime)

    file_name = time_sorted_list[len(time_sorted_list) - 1]
    return file_name

def verifyTextInFile(strvalue, filetype):
    success = 0
    file_name = get_latest_file()
    Module.cfg.file_name=file_name
    if filetype=="pdf":
        try:
            with open(file_name, 'rb') as f:
                pdfReader = PyPDF2.PdfFileReader(f)
                count = 0
                for count in range(pdfReader.numPages):
                    pageObj = pdfReader.getPage(count)
                    text = pageObj.extractText()
                    if strvalue in text:
                        success=1
                        f.close()
                        return success
                    else:
                        continue
                f.close()
        except:
                Module.logger.ERROR("Exception occurred in reading downloaded file" + file_name)
    else:
        with open(file_name, 'rt') as f:
            try:
                if filetype=="csv":
                    reader = csv.reader(f, delimiter=',')
                    for row in reader:
                        for field in row:
                            if str(field) == strvalue:
                                Module.logger.INFO("text found in file"+file_name)
                                success=1
                            if success==1:
                                f.close()
                                return success
                            else:
                                continue
                    f.close()
                elif filetype=="txt":
                    contents = f.read()
                    f.close()
                    contents = contents.replace("\n", "")
                    Module.logger.DEBUG(contents)
                    if strvalue in contents:
                        success = 1
                        Module.logger.INFO(strvalue+ " found in file" + file_name)
                        return success
                    else:
                        return success
                elif filetype=="xml":
                    with open(file_name) as f:
                        tree = ET.parse(f)
                        root = tree.getroot()
                        for elem in root.getiterator():
                            try:
                                if strvalue in str(elem.text) or strvalue in str(elem.attrib) or strvalue in str(elem.tag):
                                    success=1
                                if success==1:
                                    Module.logger.INFO(strvalue + " found in file" + file_name)
                                    f.close()
                                    return success
                                else:
                                    continue
                            except AttributeError:
                                f.close()
                                Module.logger.ERROR("Exception occurred in read downloaded file" + file_name)
                                raise
                        f.close()
            except OSError or IOError:
                Module.logger.ERROR("Exception occurred in read downloaded file" + file_name)
                raise

"""
Verify a string in email of gmail Inbox folder.
"""
def verify_email(content):
    success=0
    M = imaplib.IMAP4_SSL('imap.gmail.com')
    try:
        rv, data = M.login('mtmotomasun', 'mtmotomasun@123')
    except imaplib.IMAP4.error:
        Module.logger.ERROR("Login failed in email account")
        raise

    rv, data = M.select('Inbox')
    if rv == 'OK':
        Module.logger.INFO("Processing mailbox...\n")
        rv, data = M.search(None, "Unseen")
        if rv != 'OK':
            Module.logger.ERROR("No messages found in Inbox!")
            return success

        for num in data[0].split():
            rv, data = M.fetch(num, '(RFC822)')
            if rv != 'OK':
                Module.logger.ERROR("ERROR in fetching message from Inbox")
                return success
            msg = email.message_from_bytes(data[0][1])
            if content in str(msg):
                success==1
                return success
            else:
                continue
        M.close()
    else:
        Module.logger.ERROR("Unable to open mailbox")
        return success


