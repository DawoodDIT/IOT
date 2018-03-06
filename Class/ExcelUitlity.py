import openpyxl
import os

class ExcelUtility:

    def fnGetValueFromExcel(strExcelName, strSheetName, struniquecolname1, struniquecolvalue1, struniquecolname2, struniquecolvalue2,strValue):
        cwd = os.getcwd()
        pcwd = "\\".join(cwd.split('\\')[:-1])
        confdir = cwd +"\\Config"
        wb = openpyxl.load_workbook(confdir +"\\"+strExcelName)
        sheet = wb.get_sheet_by_name(strSheetName)
        maxrows = sheet.max_row
        maxcols = sheet.max_column
        blnfound = "false"
        for rownum1 in range(1, maxrows):
            for colnum1 in range(1, maxcols + 1):
                if sheet.cell(row=rownum1, column= colnum1).value == struniquecolname1 :
                    uniquerownumber1 = rownum1
                    uniquecolnumber1 = colnum1
                    print("Row Number for 1st unique col name is " + str(uniquerownumber1) + " And Col number for 1st Unique Col Name is " + str(uniquecolnumber1))
        for rownum2 in range(1, maxrows):
            for colnum2 in range(1, maxcols+1):
                if sheet.cell(row=rownum2, column=colnum2).value == struniquecolname2 :
                    uniquerownumber2 = rownum2
                    uniquecolnumber2 = colnum2
                    print("Row Number for 2nd unique col name  is " + str(uniquerownumber2) + " And Unique Col Name for 2nd Unique Col Name is " + str(uniquecolnumber2))
        for rownum3 in range(1, maxrows):
            if sheet.cell(row=rownum3, column=uniquecolnumber1).value == struniquecolvalue1 and sheet.cell(row=rownum3,column=uniquecolnumber2).value == struniquecolvalue2:
                uniquerownumber3 = rownum3
                print("The Unique Value is present in row number" + str(uniquerownumber3) + " and column number " +str(uniquecolnumber1))
                blnfound = "true"
                break

        if blnfound != "true":
            print("The Value is not present in the excel")
        else:
            for getcolnum in range(1, maxcols + 1):
                if sheet.cell(row=1, column=getcolnum).value == strValue:
                    print("Uniquecolvalue = " + str(getcolnum))
            c = sheet.cell(row=uniquerownumber3, column=getcolnum).value
            print("The Value of "+strValue+" based on Unique Column and Row Values is : "  + str(c))

    #fnGetValueFromExcel("example.xlsx","Sheet1","Type","Customer","Priority","Primary","MSISDN")

    def fnReadDataFromExcel(strExcelName,strSheetName,strUniqueColNames, strUniqueColValues, strValue):
        cwd = os.getcwd()
        pcwd = "\\".join(cwd.split('\\')[:-1])
        confdir = cwd + "\\Config"
        wb = openpyxl.load_workbook(confdir + "\\" + strExcelName)
        sheet = wb.get_sheet_by_name(strSheetName)
        maxrows = sheet.max_row
        maxcols = sheet.max_column
        blnfound = False
        blnList = False

        if '|'  in strUniqueColValues:
         ColValuesList = strUniqueColValues.split('|')
         blnList = True
        else:
          ColValuesList = strUniqueColValues


        for r in range(1,maxrows+1):
            counter = 0
            for c in range(1,maxcols+1):
                for i in range(0,len(ColValuesList)):
                    if sheet.cell(row=r,column=c).value == ColValuesList[i]:
                        counter+=1
                        if counter == len(ColValuesList)-1 and blnList == True:
                            UniqueRowNo = r
                            print("Unique Row number is : "+str(UniqueRowNo))
                            blnfound = True
                            break
                        if blnList!=True:
                            if counter == 1:
                                UniqueRowNo = r
                                print("Unique Row number is : " + str(UniqueRowNo))
                                blnfound = True
                                break
        if blnfound != True:
            print("The Value is not present in the excel")
        else:
            for getcolnum in range(1, maxcols + 1):
                if sheet.cell(row=1, column=getcolnum).value == strValue:
                    print("Uniquecolvalue = " + str(getcolnum))
                    break
            c = sheet.cell(row=UniqueRowNo, column=getcolnum).value
            return c
            print("The Value of " + strValue + " based on Unique Column and Row Values is : " + str(c))

        fnReadDataFromExcel("Data.xlsx", "Sheet2", "Type|Name", "text|MSISDN", "locator")